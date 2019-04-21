# -*- coding: utf-8 -*-
"""
Created on Sat Apr  6 19:47:29 2019

@author: Administrator
"""

import xml.etree.ElementTree as elemTree 

tree = elemTree.parse('test002.xml')

xmlStr = '''
<users>
    <user grade="gold">
        <name>Kim Cheol Soo</name>
        <age>25</age>
        <birthday>19940215</birthday>
    </user>
    
    <user grade="diamond">
        <name>Kim Yoo Mee</name>
        <age>21</age>
        <birthday>19980417</birthday>
    </user>
</users>
        '''
        
        
# =============================================================================
# tree = elemTree.fromstring(xmlStr)
# 
# user = tree.find('./user[2]')
# 
# 
# print(user.tag)
# print(user.attrib)
# print(user.get('grade'))
# 
# userName = user.find('name')
# userName1 = user.find('age')
# userName2 = user.find('birthday')
# 
# 
# print(userName.text +","+ userName1.text +","+ userName2.text)
# 
# =============================================================================
for user in tree.findall('./user'):print(user.tag)

############################################################

# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""



import sys
print("python 버전 : {}".format(sys.version))

import pandas as pd
print("pandas 버전 : {}".format(pd.__version__))

import matplotlib
print("matplotlib 버전 : {}".format(matplotlib.__version__))

import numpy as np
print("numpy 버전 : {}".format(np.__version__))

import scipy as sp
print("scipy 버전 : {}".format(sp.__version__))

import IPython
print("IPython 버전 : {}".format(IPython.__version__))

import sklearn
print("sklearn : {}".format(sklearn.__version__))

data=list(range(1,11))

print(data)

# %matplotlib inline
import matplotlib.pyplot as plt
plt.plot(data,'r')

############################################################

from openpyxl import load_workbook

# C:\Users\Administrator\Documents\testFile

load_wb = load_workbook("/Users/Administrator/Documents/testFile/전국cctv표준데이터_3.xlsx", data_only=True)

#시트 이름으로 불러오기
load_ws = load_wb['Sheet1']

print(load_ws['a1'].value)

print(load_ws.cell(1,2).value)



